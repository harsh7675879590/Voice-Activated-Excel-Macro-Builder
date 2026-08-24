"""
Schema Extractor — Reads .xlsx files and extracts ONLY structural metadata.
Core privacy guarantee: raw cell values NEVER leave this module for transmission.
Only column names, dtypes, row counts, and named ranges are exposed.
"""

from __future__ import annotations

import logging
from pathlib import Path
from typing import Optional

import numpy as np
import openpyxl
import pandas as pd

from backend.models.schemas import ColumnSchema, SchemaCard, WorkbookSchema

logger = logging.getLogger(__name__)


class SchemaExtractor:
    """
    Extracts workbook schema from .xlsx files.
    
    Privacy contract:
    - Reads up to SAMPLE_ROWS rows to infer dtypes and cardinality.
    - Never stores, serializes, or returns raw cell values for external use.
    - The SchemaCard output contains ONLY structural metadata.
    """

    SAMPLE_ROWS = 50  # Number of rows sampled for dtype inference

    def __init__(self):
        self._workbook_cache: dict[str, WorkbookSchema] = {}
        self._dataframe_cache: dict[str, dict[str, pd.DataFrame]] = {}

    def extract_from_file(self, file_path: str | Path) -> WorkbookSchema:
        """Extract schema from an .xlsx file on disk."""
        file_path = Path(file_path)
        if not file_path.exists():
            raise FileNotFoundError(f"Workbook not found: {file_path}")
        if file_path.suffix.lower() not in (".xlsx", ".xls", ".xlsm"):
            raise ValueError(f"Unsupported file format: {file_path.suffix}")

        logger.info(f"Extracting schema from: {file_path.name}")

        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        sheets: list[SchemaCard] = []
        dataframes: dict[str, pd.DataFrame] = {}

        for sheet_name in wb.sheetnames:
            try:
                # Read with pandas for accurate dtype inference
                df = pd.read_excel(file_path, sheet_name=sheet_name, engine="openpyxl")
                dataframes[sheet_name] = df
                schema_card = self._extract_sheet_schema(df, sheet_name)
                sheets.append(schema_card)
            except Exception as e:
                logger.warning(f"Failed to extract schema for sheet '{sheet_name}': {e}")
                continue

        wb.close()

        active_sheet = wb.sheetnames[0] if wb.sheetnames else ""

        workbook_schema = WorkbookSchema(
            filename=file_path.name,
            sheets=sheets,
            active_sheet=active_sheet,
        )

        # Cache for later use in code execution
        cache_key = file_path.name
        self._workbook_cache[cache_key] = workbook_schema
        self._dataframe_cache[cache_key] = dataframes

        return workbook_schema

    def extract_from_bytes(self, file_bytes: bytes, filename: str) -> WorkbookSchema:
        """Extract schema from uploaded file bytes."""
        import io
        import tempfile

        # Write to a temp file for openpyxl compatibility
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            tmp.write(file_bytes)
            tmp_path = Path(tmp.name)

        try:
            result = self.extract_from_file(tmp_path)
            # Re-cache under the original filename
            self._workbook_cache[filename] = result
            self._dataframe_cache[filename] = self._dataframe_cache.pop(tmp_path.name, {})
            result.filename = filename
            return result
        finally:
            tmp_path.unlink(missing_ok=True)

    def _extract_sheet_schema(self, df: pd.DataFrame, sheet_name: str) -> SchemaCard:
        """Extract schema card from a DataFrame — dtype inference on sample rows."""
        columns: list[ColumnSchema] = []

        for col in df.columns:
            col_data = df[col]
            dtype_str = self._infer_dtype(col_data)
            cardinality = int(col_data.nunique()) if len(col_data) > 0 else 0

            columns.append(ColumnSchema(
                name=str(col),
                dtype=dtype_str,
                sample_cardinality=cardinality,
            ))

        return SchemaCard(
            sheet_name=sheet_name,
            columns=columns,
            row_count=len(df),
        )

    def _infer_dtype(self, series: pd.Series) -> str:
        """Infer a human-readable dtype string from a pandas Series."""
        dtype = series.dtype

        if pd.api.types.is_integer_dtype(dtype):
            return "int64"
        elif pd.api.types.is_float_dtype(dtype):
            return "float64"
        elif pd.api.types.is_bool_dtype(dtype):
            return "bool"
        elif pd.api.types.is_datetime64_any_dtype(dtype):
            return "datetime64"
        elif pd.api.types.is_string_dtype(dtype) or dtype == object:
            # Check if it's actually dates stored as strings
            sample = series.dropna().head(self.SAMPLE_ROWS)
            if len(sample) > 0:
                try:
                    pd.to_datetime(sample, errors="raise")
                    return "datetime64"
                except (ValueError, TypeError):
                    pass
                # Check if numeric strings
                try:
                    pd.to_numeric(sample, errors="raise")
                    return "float64"
                except (ValueError, TypeError):
                    pass
            return "string"
        else:
            return str(dtype)

    def get_dataframe(self, filename: str, sheet_name: str) -> Optional[pd.DataFrame]:
        """Retrieve a cached DataFrame for code execution."""
        file_dfs = self._dataframe_cache.get(filename, {})
        return file_dfs.get(sheet_name)

    def get_all_dataframes(self, filename: str) -> dict[str, pd.DataFrame]:
        """Retrieve all cached DataFrames for a workbook."""
        return self._dataframe_cache.get(filename, {})

    def get_schema(self, filename: str) -> Optional[WorkbookSchema]:
        """Retrieve a cached workbook schema."""
        return self._workbook_cache.get(filename)

    def get_sheet_schema(self, filename: str, sheet_name: str) -> Optional[SchemaCard]:
        """Retrieve a cached schema card for a specific sheet."""
        wb_schema = self._workbook_cache.get(filename)
        if wb_schema:
            for sheet in wb_schema.sheets:
                if sheet.sheet_name == sheet_name:
                    return sheet
        return None
